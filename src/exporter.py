"""
Excel export and formatting module for Hotel Sales Analytics.

Handles exporting analysis results to Excel with professional formatting
including frozen headers, auto-width columns, and number formatting.

Author: Hotel Sales Analytics Team
Version: 1.0.0
"""

import logging
import pandas as pd
from pathlib import Path
from typing import Dict, Optional
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from config import (
    OUTPUT_FILE, OUTPUT_DIR, DETAIL_SHEET_NAME, SUMMARY_SHEET_NAME,
    SUMMARY_LABELS, CURRENCY_FORMAT, INTEGER_FORMAT
)

logger = logging.getLogger(__name__)


# ============================================================================
# EXCEL EXPORTER
# ============================================================================

class ExcelExporter:
    """
    Handles exporting analysis results to Excel with professional formatting.

    Features:
    - Frozen header rows
    - Auto-adjusted column widths
    - Number formatting (currency, integer)
    - Professional styling (bold headers, colors, borders)
    - Excel table formatting
    - Data filters
    """

    def __init__(self, output_file: Path = OUTPUT_FILE):
        """
        Initialize ExcelExporter.

        Args:
            output_file (Path): Output Excel file path
        """
        self.output_file = output_file
        self.logger = logging.getLogger(self.__class__.__name__)

        # Create output directory if it doesn't exist
        self.output_file.parent.mkdir(parents=True, exist_ok=True)

    def export_analysis(self, analysis_df: pd.DataFrame,
                       summary_stats: Dict) -> bool:
        """
        Export analysis results to Excel with formatting.

        Args:
            analysis_df (pd.DataFrame): Analysis results
            summary_stats (dict): Summary statistics

        Returns:
            bool: True if export successful, False otherwise
        """
        try:
            self.logger.info(f"Exporting to {self.output_file}")

            # Create Excel writer
            with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
                # Write detail analysis sheet
                analysis_df.to_excel(
                    writer,
                    sheet_name=DETAIL_SHEET_NAME,
                    index=False
                )

                # Write summary sheet
                summary_df = self._create_summary_sheet(summary_stats)
                summary_df.to_excel(
                    writer,
                    sheet_name=SUMMARY_SHEET_NAME,
                    index=False,
                    header=False
                )

            # Apply formatting
            self._format_workbook(self.output_file)

            self.logger.info(f"Successfully exported to {self.output_file}")
            return True

        except Exception as e:
            self.logger.error(f"Error exporting to Excel: {type(e).__name__}: {str(e)}")
            return False

    def _create_summary_sheet(self, summary_stats: Dict) -> pd.DataFrame:
        """
        Create summary sheet data.

        Args:
            summary_stats (dict): Summary statistics

        Returns:
            pd.DataFrame: Summary data formatted for Excel
        """
        summary_data = []

        # Classification Summary section
        summary_data.append(["CLASSIFICATION SUMMARY", ""])
        summary_data.append([SUMMARY_LABELS['total_analyzed'],
                           summary_stats['total_analyzed']])
        summary_data.append([SUMMARY_LABELS['total_repeater'],
                           summary_stats['total_repeater']])
        summary_data.append([SUMMARY_LABELS['total_sleeping'],
                           summary_stats['total_sleeping']])
        summary_data.append([SUMMARY_LABELS['total_newcomer'],
                           summary_stats['total_newcomer']])

        # Empty row
        summary_data.append(["", ""])

        # Revenue Opportunity section
        summary_data.append(["REVENUE OPPORTUNITY", ""])
        summary_data.append([SUMMARY_LABELS['sleeping_room_nights'],
                           summary_stats['sleeping_room_nights']])
        summary_data.append([SUMMARY_LABELS['sleeping_revenue'],
                           summary_stats['sleeping_revenue']])

        return pd.DataFrame(summary_data, columns=["Metric", "Value"])

    def _format_workbook(self, file_path: Path):
        """
        Apply professional formatting to workbook.

        Args:
            file_path (Path): Path to Excel file
        """
        try:
            wb = load_workbook(file_path)

            # Format detail sheet
            if DETAIL_SHEET_NAME in wb.sheetnames:
                self._format_detail_sheet(wb[DETAIL_SHEET_NAME])

            # Format summary sheet
            if SUMMARY_SHEET_NAME in wb.sheetnames:
                self._format_summary_sheet(wb[SUMMARY_SHEET_NAME])

            wb.save(file_path)
            self.logger.info("Workbook formatting completed")

        except Exception as e:
            self.logger.error(f"Error formatting workbook: {type(e).__name__}: {str(e)}")

    def _format_detail_sheet(self, ws):
        """
        Format detail analysis sheet.

        Args:
            ws: Worksheet object
        """
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4",
                                 fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Format header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Format data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Format numeric columns
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            # Last Visit Year column (D)
            if row[2].value is not None:
                row[2].alignment = Alignment(horizontal="center", vertical="center")

            # Last Room Night column (E) - Integer format
            if row[3].value is not None:
                row[3].number_format = INTEGER_FORMAT
                row[3].alignment = Alignment(horizontal="right", vertical="center")

            # Last Revenue column (F) - Currency format
            if row[4].value is not None:
                row[4].number_format = CURRENCY_FORMAT
                row[4].alignment = Alignment(horizontal="right", vertical="center")

        # Freeze panes (freeze header row)
        ws.freeze_panes = "A2"

        # Auto-adjust column widths
        self._auto_adjust_columns(ws)

        # Add auto-filter
        ws.auto_filter.ref = f"A1:E{ws.max_row}"

    def _format_summary_sheet(self, ws):
        """
        Format summary sheet.

        Args:
            ws: Worksheet object
        """
        # Define styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4",
                                 fill_type="solid")
        section_font = Font(bold=True, size=11, color="FFFFFF")
        section_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5",
                                   fill_type="solid")

        label_font = Font(size=11)
        value_font = Font(size=11, bold=True)

        alignment_center = Alignment(horizontal="center", vertical="center")
        alignment_right = Alignment(horizontal="right", vertical="center")

        # Format cells
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row),
                                      start=1):
            # Section headers (Classification Summary, Revenue Opportunity)
            if row[0].value in ["CLASSIFICATION SUMMARY", "REVENUE OPPORTUNITY"]:
                row[0].font = section_font
                row[0].fill = section_fill
                row[0].alignment = alignment_center
                row[1].fill = section_fill

            # Data rows
            elif row[0].value:
                row[0].font = label_font
                row[1].font = value_font
                row[1].alignment = alignment_right

                # Format revenue values as currency
                if "Revenue" in str(row[0].value):
                    row[1].number_format = CURRENCY_FORMAT

                # Format room nights as integer
                if "Room Nights" in str(row[0].value):
                    row[1].number_format = INTEGER_FORMAT

        # Set column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20

    def _auto_adjust_columns(self, ws):
        """
        Auto-adjust column widths based on content.

        Args:
            ws: Worksheet object
        """
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            # Set column width with some padding
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width


# ============================================================================
# EXPORT UTILITIES
# ============================================================================

def export_to_excel(analysis_df: pd.DataFrame, summary_stats: Dict,
                   output_file: Path = OUTPUT_FILE) -> bool:
    """
    Convenience function to export analysis results to Excel.

    Args:
        analysis_df (pd.DataFrame): Analysis results
        summary_stats (dict): Summary statistics
        output_file (Path): Output file path

    Returns:
        bool: True if successful, False otherwise
    """
    exporter = ExcelExporter(output_file)
    return exporter.export_analysis(analysis_df, summary_stats)
